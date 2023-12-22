import streamlit as st
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.naive_bayes import MultinomialNB
from sklearn.metrics import accuracy_score
import pickle
import os
import datetime
from io import BytesIO
import io
from io import StringIO
import base64
import xlsxwriter
import time
from openpyxl import load_workbook


st.set_page_config(
    page_title="Tax Package ML Classification Model",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': 'mailto:miguel.karim@karimortega.com'
    }
)



st.image("https://www.kellanovaus.com/content/dam/NorthAmerica/kellanova-us/images/logo.svg", width=120)
# st.header('Machine Learnig Model')
st.subheader('Tax Package - Related Party Operations Category Classification Machine Learning Model')

# st.divider()

start_time01 = time.time()
uploaded_FBL3N_train = st.sidebar.file_uploader("Upload FBL3N file which contains historical data classified to train the Machine Learning Model", type=["xlsx"], accept_multiple_files=False)
st.sidebar.divider()
uploaded_new_FBL3N = st.sidebar.file_uploader("Upload the file which contains the new dataset to be classified", key="new_FBL3N", type=["xlsx"], accept_multiple_files=False)
uploaded_masters = st.sidebar.file_uploader("Upload masterdata file which contains the Chart of Accounts and Subcodes", key="masters", type=["xlsx"], accept_multiple_files=False)
st.sidebar.divider()
if uploaded_FBL3N_train and uploaded_new_FBL3N and uploaded_masters:
    FBL3N_full = pd.read_excel(uploaded_FBL3N_train, engine='openpyxl', sheet_name='FBL3N', 
                               dtype = {'Subcode': str, 'Company Code': str, 'Document Type': str, 'Account': str, 'Text': str,
                                        'Reference': str, 'Document Header Text': str, 'User Name': str, 'Tax Code': str,})
    FBL3N_new = pd.read_excel(uploaded_new_FBL3N, engine='openpyxl', sheet_name='FBL3N',
                dtype = {'Subcode': str, 'Company Code': str, 'Document Type': str, 'Account': str, 'Document Number': str,
                        'Text': str, 'Reference': str, 'Document Header Text': str, 'User Name': str, 'Tax Code': str,})
    accounts = pd.read_excel(uploaded_masters, engine='openpyxl', sheet_name='GL_Accounts',
                dtype = {'GL_Account': str, 'Description': str, 'Country': str, 'CoCd': str})
    subcodes = pd.read_excel(uploaded_masters, engine='openpyxl', sheet_name='Subcodes',
                  dtype={'Code_Type': str, 'Code': str, 'Code_Desc': str, 'Code_Type_RP': str,
                         'Code_RP': str, 'Code_Desc_RP': str,})


    

    # Paso 2: Rellenar las celdas "NaN" como celdas vacías ('') en las columnas especificadas
    columnas_rellenar = ['Company Code', 'Document Type', 'Account', 'Text', 'Reference', 'Document Header Text', 'User Name', 'Tax Code']
    FBL3N_full[columnas_rellenar] = FBL3N_full[columnas_rellenar].fillna('')
    # FBL3N_full.dropna(subset=columnas_rellenar, how='any', inplace=True)


    # FBL3N_full
    # Paso 3: Crear una nueva columna 'ML' con el contenido de las columnas especificadas
    FBL3N_full['CONCAT'] = FBL3N_full['Company Code'] + (FBL3N_full['Document Number'].astype(str))
    FBL3N_full['ML'] = FBL3N_full['Company Code'] + ' ' + FBL3N_full['Document Type'] + ' ' + FBL3N_full['Account'] + ' ' + FBL3N_full['Text'] + ' ' + FBL3N_full['Reference'] + ' ' + FBL3N_full['Document Header Text'] + ' ' + FBL3N_full['User Name'] + ' ' + FBL3N_full['Tax Code']
    FBL3N_full['Id'] = FBL3N_full['Company Code'] + ' ' + FBL3N_full['Document Type'] + ' ' + (FBL3N_full['Document Number'].astype(str)) + ' ' + (FBL3N_full['Amount in doc. curr.'].astype(str)) + ' ' + (FBL3N_full['Posting Date'].astype(str))

    #---------- Subcode_td: Columna que contiene el Subcode de train data para posteriormente cruzar con el dataset clasificado
    FBL3N_full['Subcode_td'] = FBL3N_full['Company Code'] + (FBL3N_full['Document Number'].astype(str)) + FBL3N_full['Document Type'] + (FBL3N_full['Posting period'].astype(str)) + (FBL3N_full['Amount in doc. curr.'].astype(str))
    # st.divider()
    st.caption('ML FBL3N train dataset')
    # Revisión de los subcodigos asignados para poder mostrar el texto no estandarizado
    # subcodes_unique = FBL3N_full['Subcode'].unique()
    # subcodes_options = st.multiselect('Selecciona la clasificación para filtar el dataframe', subcodes_unique, subcodes_unique)
    # FBL3N_filtered = FBL3N_full[FBL3N_full['Subcode'].isin(subcodes_options)]
    # st.dataframe(FBL3N_filtered)
    # Mostrar el dataframe sin filtrar
    st.dataframe(FBL3N_full)
    st.divider()

    
    FBL3N_previous_subcode = FBL3N_full[['Subcode_td', 'Subcode']].drop_duplicates()
    FBL3N_previous_subcode["conteo"] = FBL3N_previous_subcode.groupby('Subcode_td')['Subcode_td'].transform("size")
    FBL3N_previous_subcode.rename(columns={'Subcode': 'Subcode_assigned'}, inplace=True)
    st.caption('Registros unicos')
    st.dataframe(FBL3N_previous_subcode)
    FBL3N_train = FBL3N_full[['ML', 'Subcode']].drop_duplicates()
    # FBL3N_train

    X = FBL3N_train['ML']
    y = FBL3N_train['Subcode']
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    # Vectorizar los datos de texto utilizando TF-IDF
    tfidf_vectorizer = TfidfVectorizer(max_features=1000)
    X_train_tfidf = tfidf_vectorizer.fit_transform(X_train)
    X_test_tfidf = tfidf_vectorizer.transform(X_test)

    # Entrenar un modelo de clasificación
    modelo = MultinomialNB()
    modelo.fit(X_train_tfidf, y_train)

    # Realizar predicciones en el conjunto de prueba
    y_pred = modelo.predict(X_test_tfidf)

    # Calcular la precisión del modelo en el conjunto de prueba
    accuracy = accuracy_score(y_test, y_pred)
    accuracy = "{:.2%}".format(accuracy)
    # print(accuracy)
    # st.caption('El modelo de aprendizaje finalizó y una vez que el modelo fue probado, dio un porcentaje de accuracy del:')
    st.metric(label="Model Accuracy", value=accuracy, delta=accuracy)
    
    end_time01 = time.time()
    processing_time01 = end_time01 - start_time01
    processing_time_formatted01 = "{:.2f}".format(processing_time01)
    st.info(f'Machine Learning model training time: {processing_time_formatted01} seconds')

    st.divider()
# st.subheader('Una vez entrenado el modelo de ML, se realizará la clasificación en el nuevo conjunto de datos')

    start_time02 = time.time()
# st.sidebar.divider()
# uploaded_new_FBL3N = st.sidebar.file_uploader("Upload the file which contains the new dataset to be classified", key="new_FBL3N", type=["xlsx"], accept_multiple_files=False)
# uploaded_masters = st.sidebar.file_uploader("Upload masterdata file which contains the Chart of Accounts and Subcodes", key="masters", type=["xlsx"], accept_multiple_files=False)
# st.sidebar.divider()
# if uploaded_new_FBL3N and uploaded_masters:
    # FBL3N_new = pd.read_excel(uploaded_new_FBL3N, engine='openpyxl', sheet_name='FBL3N',
    #             dtype = {'Subcode': str, 'Company Code': str, 'Document Type': str, 'Account': str,
    #                     'Text': str, 'Document Header Text': str, 'User Name': str,
    #                     'Tax Code': str,})
    # accounts = pd.read_excel(uploaded_masters, engine='openpyxl', sheet_name='GL_Accounts',
    #             dtype = {'GL_Account': str, 'Description': str, 'Country': str, 'CoCd': str})
    # subcodes = pd.read_excel(uploaded_masters, engine='openpyxl', sheet_name='Subcodes',
    #               dtype={'Code_Type': str, 'Code': str, 'Code_Desc': str, 'Code_Type_RP': str,
    #                      'Code_RP': str, 'Code_Desc_RP': str,})
    
    

    # Paso 2: Rellenar las celdas "NaN" como celdas vacías ('') en las columnas especificadas
    columnas_rellenar_real = ['Company Code', 'Document Type', 'Account', 'Text', 'Reference', 'Document Header Text', 'User Name', 'Tax Code']
    FBL3N_new[columnas_rellenar_real] = FBL3N_new[columnas_rellenar_real].fillna('')
    FBL3N_new['CONCAT'] = FBL3N_new['Company Code'] + (FBL3N_new['Document Number'].astype(str))
    FBL3N_new['ML'] = FBL3N_new['Company Code'] + ' ' + FBL3N_new['Document Type'] + ' ' + FBL3N_new['Account'] + ' ' + FBL3N_new['Text'] + ' ' + FBL3N_new['Reference'] + ' ' + FBL3N_new['Document Header Text'] + ' ' + FBL3N_new['User Name'] + ' ' + FBL3N_new['Tax Code']
    # FBL3N_new['Id'] = FBL3N_new['Company Code'] + ' ' + FBL3N_new['Document Type'] + ' ' + (FBL3N_new['Document Number'].astype(str)) + ' ' + (FBL3N_new['Amount in doc. curr.'].astype(str)) + ' ' + (FBL3N_new['Posting Date'].astype(str))
    # FBL3N_new['Subcode_td_1'] = FBL3N_new['Company Code'] + ' ' + (FBL3N_new['Document Number'].astype(str)) + ' ' + FBL3N_new['Document Type'] + ' ' + (FBL3N_new['Posting period'].astype(str)) + ' ' + (FBL3N_new['Amount in doc. curr.'].astype(str))
    FBL3N_new['Subcode_td_1'] = FBL3N_new['Company Code'] + (FBL3N_new['Document Number'].astype(str)) + FBL3N_new['Document Type'] + (FBL3N_new['Posting period'].astype(str)) + (FBL3N_new['Amount in doc. curr.'].astype(str))
    X_new_data_tfidf = tfidf_vectorizer.transform(FBL3N_new['ML'])
    # Realizar predicciones con el modelo entrenado en el conjunto de datos real
    FBL3N_new['Subcode_ML'] = modelo.predict(X_new_data_tfidf)

    # y_test_pred = modelo.predict(X_test_tfidf)
    # accuracy_test = accuracy_score(y_test, y_test_pred)

    FBL3N_new = FBL3N_new.merge(accounts, left_on="Account", right_on='GL_Account', how='left')
    FBL3N_new = FBL3N_new.merge(subcodes, left_on="Subcode_ML", right_on='Code', how='left')
    #---------------Funciones para subcodes fijas-------------------
    def sc_121_1(row):
        if row['Reference'].startswith("00015-") and row['Document Header Text'].startswith("1176"):
            return "121"
        else:
            return ''

    def sc_121_2(row):
        if row['Reference'].startswith("00016-") and (row['Document Header Text'].startswith("1176") or row['Document Header Text'].startswith("8")):
            return "121"
        else:
            return ''

    def sc_121_3(row):
        if row['Reference'].startswith("1176") and row['Document Type'].startswith("RV"):
            return "121"
        else:
            return ''

    def sc_221_1(row):
        if row['Reference'].startswith("1176") and row['Document Type'].startswith("RN"):
            return "221"
        else:
            return ''

    def sc_221_2(row):
        if row['Reference'].startswith("CR"):
            return "221"
        else:
            return ''

    
    def sc_150(row):
    # Verificar las condiciones
        if "loan int" in str(row['Document Header Text']).lower() and row['Reference'].startswith(str(row['Company Code']) + str(row['CoCd'])):
            return "150"
        else:
            return ''

    def sc_250(row):
    # Verificar las condiciones
        if "loan int" in str(row['Document Header Text']).lower() and row['Reference'].startswith(str(row['CoCd']) + str(row['Company Code'])):
            return "250"
        else:
            return ''

    def sc_300_1(row):
    # Verificar las condiciones
        if "wf-batch" in str(row['User Name']).lower():
            return "300"
        else:
            return ''

    
    FBL3N_new['SC_1'] = FBL3N_new.apply(sc_121_1, axis=1)
    FBL3N_new['SC_2'] = FBL3N_new.apply(sc_121_2, axis=1)
    FBL3N_new['SC_3'] = FBL3N_new.apply(sc_121_3, axis=1)
    FBL3N_new['SC_4'] = FBL3N_new.apply(sc_221_1, axis=1)
    FBL3N_new['SC_5'] = FBL3N_new.apply(sc_221_2, axis=1)
    FBL3N_new['SC_6'] = FBL3N_new.apply(sc_150, axis=1)
    FBL3N_new['SC_7'] = FBL3N_new.apply(sc_250, axis=1)
    FBL3N_new['SC_8'] = FBL3N_new.apply(sc_300_1, axis=1)
    FBL3N_new['SC_concat'] = FBL3N_new['SC_1'] + FBL3N_new['SC_2'] + FBL3N_new['SC_3'] + FBL3N_new['SC_4'] + FBL3N_new['SC_5'] + FBL3N_new['SC_6'] + FBL3N_new['SC_7'] + FBL3N_new['SC_8']

    def Subcode_Correction(row):
    # Verificar las condiciones
        if row['SC_concat'] != '' and (row['SC_concat'] != row['Subcode_ML'] ):
            return row['SC_concat']
        else:
            return row['Subcode_ML']
    FBL3N_new['SC_Fix'] = FBL3N_new.apply(Subcode_Correction, axis=1)



    
    FBL3N_summary = FBL3N_new.copy()
    FBL3N_summary['K1'] = FBL3N_summary['Company Code'] + FBL3N_summary['CoCd'] + FBL3N_summary['Document currency'] + (FBL3N_summary['Subcode_ML'].astype(str))
    FBL3N_summary['K2'] = FBL3N_summary['CoCd'] + FBL3N_summary['Company Code'] + FBL3N_summary['Document currency'] + (FBL3N_summary['Code_RP'].astype(str))
    FBL3N_summary = FBL3N_summary.groupby(by=['Company Code', 'CoCd', 'Subcode_ML', 'Code_Type', 'Code_Desc', 'Code_RP', 'Document currency', 'K1', 'K2'], as_index=False)['Amount in doc. curr.'].sum()
   
    FBL3N_summary2 = FBL3N_summary.copy()
    FBL3N_summary2.columns = [col_name + '_k2' for col_name in FBL3N_summary2]
    FBL3N_summary = FBL3N_summary.merge(FBL3N_summary2, left_on="K1", right_on='K2_k2', how='left')
    FBL3N_summary = FBL3N_summary.drop(columns=['K1', 'K2','Company Code_k2','CoCd_k2','Subcode_ML_k2','Code_Type_k2','Code_Desc_k2','Code_RP_k2','K1_k2','K2_k2'])
    # FBL3N_summary = FBL3N_summary[['Document currency', 'Amount in doc. curr.', 'Document currency_k2', 'Amount in doc. curr._k2']].fillna(0)
    # FBL3N_summary = FBL3N_summary[['Amount in doc. curr.', 'Amount in doc. curr._k2']].fillna(0)
    FBL3N_summary['Diferencia'] = FBL3N_summary['Amount in doc. curr.'] + FBL3N_summary['Amount in doc. curr._k2']
    

    reg_clas = len(FBL3N_new[FBL3N_new['SC_concat'] != ''])
    st.write(reg_clas)
    tab1, tab2 = st.tabs(["Resumen", "Detalle"])
    
    with tab1:
        st.subheader(f'Resumen')
        st.dataframe(FBL3N_summary)
        st.write(FBL3N_summary.columns)

    with tab2:
        FBL3N_new = FBL3N_new.merge(FBL3N_previous_subcode, left_on="Subcode_td_1", right_on='Subcode_td', how='left')
        # FBL3N_new['Key1'] = FBL3N_new['Company Code'] + FBL3N_new['CoCd'] + (FBL3N_new['Document Date'].astype(str)) + (FBL3N_new['Amount in doc. curr.'].astype(str))
        # FBL3N_new['Key2'] = FBL3N_new['CoCd'] + FBL3N_new['Company Code'] + (FBL3N_new['Document Date'].astype(str)) + (-FBL3N_new['Amount in doc. curr.']).astype(str)
        
        # FBL3N_new['Counter1'] = FBL3N_new.groupby('Key1').cumcount()
        # FBL3N_new['Counter1'] += 0 # Sumar 1 al contador para que comience desde 1 en lugar de 0
        # FBL3N_new['Key_1'] = FBL3N_new['Key1'] + FBL3N_new['Counter1'].astype(str) # Crear una nueva columna 'key_modified' que contiene la columna 'key' con el contador
        # FBL3N_new['Counter2'] = FBL3N_new.groupby('Key2').cumcount()
        # FBL3N_new['Counter2'] += 0 # Contador para que comience desde 0
        # FBL3N_new['Key_2'] = FBL3N_new['Key2'] + FBL3N_new['Counter2'].astype(str) # Crear una nueva columna 'key_modified' que contiene la columna 'key' con el contador
        
        FBL3N_real2 = FBL3N_new.copy()
        FBL3N_real2.columns = [col_name + '_k2' for col_name in FBL3N_real2]
        # FBL3N_real = FBL3N_real.merge(FBL3N_real2, left_on="Key1", right_on='Key2_k2', how='left')
        # st.dataframe(FBL3N_tobe_class)
        # st.dataframe(FBL3N_classified)
        st.dataframe(FBL3N_new)
    
    end_time02 = time.time()
    processing_time02 = end_time02 - start_time02
    processing_time_formatted02 = "{:.4f}".format(processing_time02)
    st.info(f'Subcodes has been assigned to the new FBL3N dataset according to the Machine Learning Model in: {processing_time_formatted02} seconds')

#---------------abrir archivo de excel y sobreescribir data

    excel_buffer = BytesIO()
    
    # Write the DataFrame to the Excel buffer
    FBL3N_new.to_excel(excel_buffer, index=False, sheet_name='FBL3N')
    
    # Load the Excel file template.xlsx
    excel_file_path = "Template FBL3N.xlsx"
    workbook = load_workbook(excel_file_path)
    
    # Access the specified sheet
    sheet = workbook['FBL3N']
    
    # Clear existing data in the table
    for row in sheet['A1:AF200000']:  # Assuming your table starts from A2 and goes up to M10000
        for cell in row:
            cell.value = None
    
    # Load the data from the Excel buffer into the table starting from A2
    excel_buffer.seek(0)
    sheet['A1'].value = excel_buffer
    
    # Save the modified workbook
    workbook.save(excel_file_path)
    
    # Create a download button
    st.download_button(
        label="Download Updated Excel",
        data=excel_buffer.getvalue(),
        file_name='updated_template.xlsx',
        key='download_button'
    )





    
# #-------------------------sobreescribir rchivo de excel-------------------------
#     # Replace table data with new DataFrame
    
#     sheet[table_range[0]:table_range[1]] = FBL3N_new.values
    
#     # Resize table to fit new data
#     sheet.tables["tbl_FBL3N"].ref = f"{table_range[0]}:{sheet.max_row}{sheet.max_column}"
    
#     # Create download button with modified code
#     excel_buffer = BytesIO()
#     workbook.save(excel_buffer)  # Save directly to BytesIO
#     st.download_button(
#         label="Descargar Excel",
#         data=excel_buffer.getvalue(),
#         file_name="FBL3N.xlsx",  # Updated file name
#         key="download_button",
#     )
# #-------------------------

    
# #     excel_buffer = BytesIO()

# # # Utilizar el método to_excel() pero guardar en el objeto BytesIO en lugar de un archivo local
# #     FBL3N_new.to_excel(excel_buffer, index=False, sheet_name='FBL3N')  # Asegúrate de cambiar 'Hoja1' al nombre real de tu hoja

# # # Descargar el archivo Excel en Streamlit
# #     st.download_button(
# #         label="Descargar Excel",
# #         data=excel_buffer.getvalue(),
# #         file_name='template.xlsx',  # Puedes cambiar el nombre del archivo según tus necesidades
# #         key='download_button'
# #     )





# #-------------codigo anterior---------------------
#     # # if st.checkbox("Generar Archivo de Excel"):
#     # #     start_time03 = time.time()
#     # #     fecha_actual = datetime.datetime.now()
#     # #     formato = "%Y%m%d %H%M%S"  # Formato: Año-Mes-Día Hora-Minuto-Segundo
#     # #     fecha_formateada = fecha_actual.strftime(formato)
        
#     #     # Crear un objeto de Pandas ExcelWriter y un buffer
#     # buffer = pd.ExcelWriter("output.xlsx", engine='openpyxl')

#     # # Escribir el DataFrame en la hoja de Excel
#     # FBL3N_new.to_excel(buffer, sheet_name='FBL3N', index=False)
#     # # FBL3N_real2.to_excel(buffer, sheet_name='FBL3N_2', index=False)
#     # # Cerrar el Pandas Excel writer
#     # buffer.close()

#     # # Cargar el archivo Excel en un búfer base64 y crear un enlace de descarga
#     # with open("output.xlsx", "rb") as excel_file:
#     #     b64 = base64.b64encode(excel_file.read()).decode()

#     # # Crear un enlace de descarga para el archivo Excel
#     # href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="FBL3N_{fecha_formateada}.xlsx">Download Excel File</a>'

#     # # Mostrar el enlace en Streamlit
#     # st.markdown(href, unsafe_allow_html=True)
#     #     # end_time03 = time.time()
#     #     # processing_time03 = end_time03 - start_time03
#     #     # processing_time_formatted03 = "{:.4f}".format(processing_time03)
#     #     # st.success(f'Archivo de Excel generado en un tiempo total de: {processing_time_formatted03} segundos')
